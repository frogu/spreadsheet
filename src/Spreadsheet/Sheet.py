# -*- coding: utf-8 -*-
# vim: set fileencoding=utf-8 :
'''
Created on 2010-11-18

@author: lukasz.proszek@selvita.com
'''

from functools import total_ordering
from openpyxl.reader.excel import load_workbook as xlsx_load_wb
from openpyxl.writer.excel import save_workbook as xlsx_save_wb
from sys import stderr
from test.test_decorators import memoize
import cStringIO
import codecs
import csv
import exceptions
import functools
import openpyxl
import os
import xlrd
import xlwt
# import chardet


class UTF8Recoder:
    """
    Iterator that reads an encoded stream and reencodes the input to UTF-8
    """
    def __init__(self, f, encoding):
        self.reader = codecs.getreader(encoding)(f)

    def __iter__(self):
        return self

    def next(self):
        return self.reader.next().encode("utf-8")


class UnicodeReader:
    """
    A CSV reader which will iterate over lines in the CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        f = UTF8Recoder(f, encoding)
        self.reader = csv.reader(f, dialect=dialect, **kwds)

    def next(self):
        row = self.reader.next()
        return [a if a else None for a in [unicode(s, "utf-8") for s in row]]

    def __iter__(self):
        return self

class UnicodeWriter:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([s.encode("utf-8") if s else "" for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)



def isnumeric(arg):
    '''
    test if arg is a number
    '''
    try:
        float(arg)
        return True
    except ValueError:
        return False


def col_idx_from_str(x):
    if isnumeric(x):
        return x
    else:
        try:
            return openpyxl.cell.column_index_from_string(x) - 1
        except openpyxl.shared.exc.ColumnStringIndexException:
            return None


@total_ordering
class Reader(object):

    @property
    @memoize
    def NROWS(self):
        """
        number of rows
        """
        return len(self.csvdata)

    @property
    @memoize
    def NCOLS(self):
        """number of columns"""
        return max([len(r) for r in self.csvdata])

    def __len__(self):
        return self.NROWS

    def __lt__(self, other):
        return self.NROWS * self.NCOLS < other.NROWS * other.NCOLS

    def __eq__(self, other):
        return (self.NROWS == other.NROWS) and (self.NCOLS == other.NCOLS)

    def __str__(self):
        if (self.csvfilename
            and self.dialect
            and self.NROWS
            and self.NCOLS):
            return "{0} ({1}; {2}x{3})".format(
                                           self.csvfilename,
                                           self.dialect,
                                           self.NROWS,
                                           self.NCOLS
                                           )
        else:
            return "No file read yet"

    def __init__(self, filename, sheet=None, encoding='utf-8'):
        self.csvheader = []
        if os.path.exists(filename):
            ext = filename.split('.')[-1]

            if ext.lower() == 'xlsx':
                wb = xlsx_load_wb(filename)
                if sheet:
                    sht = wb.get_sheet_by_name(sheet)
                    if not sht:
                        raise exceptions.NameError(
                        u'No such sheet {0}. Valid choices are: {1}'.format(
                                                    sheet,
                                                    wb.get_sheet_names()
                                                    )
                                                   )
                else:
                    msg = "No spreadsheet name given "
                    msg += "- defaulting to the first available spreadsheet.\n"
                    stderr.write(msg)
                    sht = wb.get_sheet_by_name(wb.get_sheet_names()[0])
                self.dialect = 'xlsx'

            elif ext.lower() == 'xls':
                wb = xlrd.open_workbook(filename)
                if sheet:
                    try:
                        sht = wb.sheet_by_name(sheet)
                    except xlrd.biffh.XLRDError:
                        msg = u"Sheet '{0]' doesn't exist. "
                        msg += "Valid choices are: {1}". format(
                                                            sheet,
                                                            wb.sheet_names()
                                                        )
                        raise exceptions.NameError(msg)
                else:
                    msg = "No spreadsheet name given "
                    msg += "- defaulting to the first available spreadsheet.\n"
                    stderr.write(msg)
                    sht = wb.sheet_by_index(0)
                self.dialect = 'xls'

            elif ext.lower() == 'csv':
                try:
                    dialect = csv.Sniffer().sniff(open(filename).read(4096))
                except:
                    dialect = None
                # print 'dialect', dialect.delimiter
                # print repr(dialect)
                fi = UTF8Recoder(open(filename), encoding)
                if sheet:
                    sht = [row  for row in csv.reader(
                                                      fi,
                                                      delimiter=sheet
                                                      )
                           ]
                elif dialect:
                    sht = [row  for row in csv.reader(
                                                      fi,
                                                      dialect
                                                      )
                           ]
                else:
                    sht = [row  for row in csv.reader(
                                                      fi,
                                                      delimiter=';'
                                                      )
                           ]
                self.dialect = 'csv'
                sht = [[unicode(a, 'utf-8') for a in s]for s in sht]
                # for i in sht:
                #    print "; ".join(i)
            else:
                msg = "File {0} cannot be read ".format(filename)
                msg += "(make sure that the extension is one of: "
                msg += " csv, xls, xlsx)"
                raise exceptions.IOError(msg)
            # print sht
            self.csvdata = sht
            self.csvheader = self.get_row(0)
            self.csvfilename = filename
        else:
            raise exceptions.IOError("File {0} doesn't exist".format(filename))

    def add_column(self, header):
        """adds an empty column to the data"""
        for i in xrange(self.NROWS):
            self.csvdata[i].append(header if i == 0 else '')

    def update_row(self, index, values):
        """updates row with specified values"""
        for k, v in values.iteritems():
            col_id = self.get_row(0).index(k)
            self.csvdata[index][col_id] = v

    def get_dims(self):
        """get data dimensions"""
        return (self.NROWS, self.NCOLS)

    def xrows(self, skipfirst=False, as_dict=False):
        """row generator"""
        for n in xrange(int(skipfirst), self.NROWS):
            yield self.get_row(n, as_dict=as_dict)

    def get_row(self, index, as_dict=None):
        """returns a specific row either as list or a dictionary"""
        if self.dialect == "csv":
            return {
                    k:v
                    for (k, v)
                    in zip(self.csvheader, self.csvdata[index])
                    } if as_dict else self.csvdata[index]
        elif self.dialect == 'xlsx':
            ver = openpyxl.__version__
            ver = ver.split('.')
            if ver[1] < 2:
                offset = 1
            else:
                offset = 0
            return [
                    self.csvdata.cell(
                                   column=offset + i,
                                   row=offset + index
                                   ).value
                     for i in xrange(self.NCOLS)
                     ]

        elif self.dialect == 'xls':
            return [d.value for d in self.csvdata.row(index)]

    def get_col(self, index):
        """returns a specific column"""
        if self.dialect == "csv":
            if not isnumeric(index):
                index = self.get_row(0).index(index)
            return zip(*self.csvdata)[index]
        elif self.dialect == 'xlsx':
            ver = openpyxl.__version__
            ver = ver.split('.')
            if ver[1] < 2:
                offset = 1
            else:
                offset = 0
            return [
                    self.csvdata.cell(
                                   column=offset + index,
                                   row=offset + i
                                   ).value
                    for i in xrange(self.NROWS)
                    ]
        elif self.dialect == 'xls':
            return [d.value for d in self.csvdata.col(index)]

    def get_cell(self, r, c):
        """returns a single cell value"""
        if self.dialect == "csv":
            return self.csvdata[r][c]
        elif self.dialect == 'xlsx':
            ver = openpyxl.__version__
            ver = ver.split('.')
            if ver[1] < 2:
                offset = 1
            else:
                offset = 0
            return self.csvdata.cell(column=offset + c, row=offset + r).value
        elif self.dialect == 'xls':
            return self.csvdata.cell_value(r, c)

    def __getitem__(self, args):
        try:
            if len(args) == 2:
                if args[0] is None:
                    return self.get_col(args[1])
                else:
                    return self.get_cell(args[0], args[1])
        except TypeError:
            return self.get_row(args)


class Writer():
    """Unified CVS, XLS, XLSX spreadsheet write class"""
    def __init__(self, filename, sheet=None):
        self.dialect = filename.split('.')[-1].lower()
        self.csvfilename = filename
        if os.path.exists(filename):
            print "OVERWRITING existing file: {0}".format(filename)
            os.unlink(filename)
            # msg="Appending to an existing workbook "
            # msg+="is not supproted at this time"
            # raise exceptions.NotImplementedError(msg)

        if self.dialect == 'csv':
            if sheet is not None:
                delimiter = sheet
            else:
                delimiter = ';'
            self.__opened_file = open(filename, 'wb')
            self.__writer = UnicodeWriter(
            # self.__writer = csv.writer(
                                       self.__opened_file,
                                       delimiter=delimiter,
                                       quotechar='"',
                                       quoting=csv.QUOTE_MINIMAL
                                       )

        elif self.dialect == 'xlsx':
            self.__wb = openpyxl.workbook.Workbook()
            self.__sht = self.__wb.create_sheet()
            if sheet:
                self.__sht._set_title(sheet)

        elif self.dialect == 'xls':
            self.__wb = xlwt.Workbook()
            if sheet:
                self.__sht = self.__wb.add_sheet(sheet)

            else:
                self.__sht = self.__wb.add_sheet('Sheet_R')
        else:
            raise exceptions.IOError("File {0} ".format(filename))

    def write(self, data):
        if self.dialect == "csv":
            # csvdata[0] = [r.encode('utf-8') for r in csvdata[0]]

            for r in data:
                self.__writer.writerow(r)
            self.__opened_file.close()

        if self.dialect == 'xlsx':
            ver = openpyxl.__version__
            ver = ver.split('.')
            if ver[1] < 2:
                offset = 1
            else:
                offset = 0

            for er, r in enumerate(data):
                for ec, c in enumerate(r):
                    if c:
                        self.__sht.cell(
                                        column=offset + ec,
                                        row=offset + er
                                        ).value = c
            xlsx_save_wb(self.__wb, self.csvfilename)

        if self.dialect == 'xls':
            for er, r in enumerate(data):
                for ec, c in enumerate(r):
                    if c:
                        self.__sht.write(er, ec, c)
            self.__wb.save(self.csvfilename)
